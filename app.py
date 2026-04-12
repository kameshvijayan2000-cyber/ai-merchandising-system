import streamlit as st
from modules.fabric_program import calculate_fabric_program
from modules.piece_weight import calculate_piece_weight
from modules.auto_directory import auto_directory_module
from modules.vendor_mailer import vendor_mailer_module
from modules.count_calculator import count_calculator_module

st.set_page_config(page_title="AI Merchandising System", layout="wide")

st.title("AI Merchandising Automation System")

st.sidebar.title("Navigation")

option = st.sidebar.selectbox(
    "Select Module",
    [
        "Home",
        "Fabric Program",
        "Count Calculator",
        "Costing Module",
        "T&A Generator",
        "AI Auto Directory",
        "Vendor Mail Sender",
        "Techpack Analysis"
    ]
)

# ================= HOME =================
if option == "Home":
    st.header("Welcome")
    st.write("Final Year Project - AI Integrated Merchandising System")


# ================= FABRIC PROGRAM =================
elif option == "Fabric Program":

    st.header("Fabric Program Calculator")

    # -------- ORDER INPUTS --------
    total_qty = st.number_input("Total Order Quantity", value=7200)
    extra_percent = st.number_input("Extra Quantity %", value=7.0)

    # -------- SIZE INPUT --------
    st.subheader("Size Configuration")

    size_input = st.text_input(
        "Enter Sizes (comma separated)",
        value="XS,S,M,L,XL"
    )

    sizes_list = [size.strip() for size in size_input.split(",") if size.strip()]

    # -------- PIECE WEIGHTS --------
    st.subheader("Piece Weights (Kg per piece)")

    body_weight = st.number_input("Body Piece Weight", value=0.350)
    rib_weight = st.number_input("Rib Piece Weight (Enter 0 if not applicable)", value=0.100)
    sj_weight = st.number_input("SJ Piece Weight (Enter 0 if not applicable)", value=0.003)

    # -------- PROCESS LOSS --------
    st.subheader("Process Loss %")

    body_loss = st.number_input("Body Process Loss %", value=13.0)
    rib_loss = st.number_input("Rib Process Loss %", value=10.0)
    sj_loss = st.number_input("SJ Process Loss %", value=0.0)

    # -------- COLOR INPUT --------
    st.subheader("Color & Size Ratio Entry")

    num_colors = st.number_input("Number of Colors", min_value=1, max_value=20, value=1)

    color_ratios = {}

    for i in range(int(num_colors)):

        st.markdown(f"### Color {i+1}")

        color_name = st.text_input(f"Enter Color Name {i+1}", key=f"color_{i}")

        size_ratios = {}

        for size in sizes_list:
            ratio = st.number_input(
                f"{size}",
                key=f"{size}_{i}",
                value=1
            )
            size_ratios[size] = ratio

        if color_name:
            color_ratios[color_name] = size_ratios

    # -------- CALCULATE --------
    if st.button("Generate Fabric Program"):

        if not color_ratios:
            st.warning("Please enter at least one valid color name.")
        else:
            df = calculate_fabric_program(
                total_qty,
                color_ratios,
                body_weight,
                rib_weight,
                sj_weight,
                extra_percent,
                body_loss,
                rib_loss,
                sj_loss
            )

            st.subheader("Fabric Summary (Kg)")
            st.dataframe(df)

elif option == "Count Calculator":
    count_calculator_module()

# ================= T&A =================
elif option == "T&A Generator":

    import pandas as pd
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import FormulaRule

    st.header("Time & Action Calendar")

    # -------- BASIC INFO --------
    style_name = st.text_input("Style Name")
    gsm = st.number_input("GSM", value=180)
    total_qty = st.number_input("Total Order Quantity", value=5000)
    cut_qty = int(total_qty * 1.07)

    dispatch_date = st.date_input("Dispatch Date")

    st.subheader("Editable Duration (Days for Each Activity)")

    activities = [
        "Order receipt (Buyer PO)",
        "Consumption calculation",
        "BOM generation",
        "PO issue for Fabric, trims",
        "Size set submission",
        "Size set Comments",
        "PP meeting",
        "Production planning updates and circulation",
        "Fabric in-house",
        "Trims In-house",
        "Cutting",
        "Fabrication/stitching",
        "Finishing",
        "Packing",
        "Inspection",
        "Dispatch"
    ]

    default_days = {act: 2 for act in activities}
    default_days["Fabric in-house"] = 8
    default_days["Fabrication/stitching"] = 10

    lead_time = {}
    for act in activities:
        lead_time[act] = st.number_input(
            f"{act} Duration (Days)",
            value=default_days[act],
            key=act
        )

    if st.button("Generate Live Gantt Excel"):

        schedule = []
        current_end = dispatch_date

        # -------- BACKWARD PLANNING --------
        for act in reversed(activities):
            duration = lead_time[act]
            planned_end = current_end
            planned_start = planned_end - datetime.timedelta(days=duration)

            schedule.append([
                act,
                planned_start,
                planned_end,
                duration
            ])

            current_end = planned_start

        schedule.reverse()

        df = pd.DataFrame(schedule, columns=[
            "Activity",
            "Planned Start",
            "Planned End",
            "Duration"
        ])

        # -------- CREATE EXCEL --------
        wb = Workbook()
        ws = wb.active
        ws.title = "Live Gantt Tracker"

        ws.append(["Style Name:", style_name])
        ws.append(["Dispatch Date:", dispatch_date])
        ws.append(["Today:", "=TODAY()"])
        ws.append([])

        ws.append([
            "Activity",
            "Planned Start",
            "Planned End",
            "Duration",
            "Start Offset",
            "Status"
        ])

        start_row = 6
        project_start = df["Planned Start"].min()

        for i, row in df.iterrows():
            excel_row = start_row + i

            offset = (row["Planned Start"] - project_start).days

            ws.append([
                row["Activity"],
                row["Planned Start"],
                row["Planned End"],
                row["Duration"],
                offset,
                f'=IF(B{excel_row}-TODAY()<0,"Delayed",IF(B{excel_row}-TODAY()<=10,"Risk","On Track"))'
            ])

        # -------- CONDITIONAL FORMATTING --------
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        dark_red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

        status_range = f"F{start_row}:F{start_row + len(df) - 1}"

        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$F{start_row}="Risk"'], fill=red_fill)
        )

        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$F{start_row}="Delayed"'], fill=dark_red_fill)
        )

        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'$F{start_row}="On Track"'], fill=green_fill)
        )

        # -------- CREATE GANTT CHART --------
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Live Production Gantt Chart"

        data = Reference(ws, min_col=5, max_col=5,
                         min_row=start_row, max_row=start_row + len(df) - 1)
        duration_data = Reference(ws, min_col=4, max_col=4,
                                  min_row=start_row, max_row=start_row + len(df) - 1)
        cats = Reference(ws, min_col=1,
                         min_row=start_row, max_row=start_row + len(df) - 1)

        chart.add_data(data)
        chart.add_data(duration_data)
        chart.set_categories(cats)

        chart.series[0].graphicalProperties.noFill = True

        ws.add_chart(chart, "H6")

        file_name = f"Live_Gantt_{style_name}.xlsx"
        wb.save(file_name)

        st.success("Live Gantt Excel Generated Successfully")

        with open(file_name, "rb") as f:
            st.download_button(
                "Download Live Gantt Excel",
                f,
                file_name=file_name
            )




# ================= AI AUTO DIRECTORY =================
elif option == "AI Auto Directory":
    auto_directory_module()

elif option == "Vendor Mail Sender":
    vendor_mailer_module()


# ================= TECHPACK =================
elif option == "Techpack Analysis":
    st.header("Techpack AI Analysis")
    st.write("Techpack reading and extraction will be done here.")
elif option == "Costing Module":

    st.header("Garment Costing Module")

    # ------------------ PIECE WEIGHT ------------------
    st.subheader("Piece Weight Calculation")

    garment_type = st.selectbox(
        "Select Garment Type",
        ["T-Shirt Full Sleeve", "T-Shirt Half Sleeve", "Track Pant"]
    )

    total_length = st.number_input("Total Length (cm)", value=70.0)
    gsm = st.number_input("Fabric GSM", value=180.0)
    extra_fabric = st.number_input("Extra Fabric (grams)", value=10.0)

    piece_weight = 0

    if garment_type in ["T-Shirt Full Sleeve", "T-Shirt Half Sleeve"]:
        chest = st.number_input("Chest (cm)", value=50.0)
        sleeve_length = st.number_input("Sleeve Length (cm) (Enter 0 if not applicable)", value=60.0)
        sleeve_width = st.number_input("Sleeve Width (cm) (Enter 0 if not applicable)", value=20.0)

        piece_weight = calculate_piece_weight(
            garment_type,
            total_length,
            chest,
            sleeve_length,
            sleeve_width,
            gsm,
            extra_fabric
        )
    else:
        thigh = st.number_input("Thigh (cm)", value=30.0)

        piece_weight = calculate_piece_weight(
            garment_type,
            total_length,
            thigh,
            0,
            0,
            gsm,
            extra_fabric
        )

    st.success(f"Piece Weight (grams): {piece_weight}")

    # Convert grams to Kg
    piece_weight_kg = piece_weight / 1000

    # ------------------ FABRIC PROCESS COST ------------------
    st.subheader("Fabric Process Cost (Per Kg)")

    process_items = {
        "Yarn": st.number_input("Yarn (Per Kg)", value=0.0),
        "Knitting": st.number_input("Knitting (Per Kg)", value=0.0),
        "Dyeing": st.number_input("Dyeing (Per Kg)", value=0.0),
        "Compacting": st.number_input("Compacting (Per Kg)", value=0.0),
        "Raising": st.number_input("Raising (Per Kg)", value=0.0),
        "Washing": st.number_input("Washing (Per Kg)", value=0.0),
        "Printing": st.number_input("Printing (Per Kg)", value=0.0)
    }

    st.markdown("### Add Custom Fabric Process")
    custom_process_name = st.text_input("Process Name")
    custom_process_cost = st.number_input("Process Cost (Per Kg)", value=0.0)

    if custom_process_name:
        process_items[custom_process_name] = custom_process_cost

    total_fabric_rate = sum(process_items.values())
    fabric_cost = piece_weight_kg * total_fabric_rate

    st.info(f"Total Fabric Rate Per Kg: ₹ {round(total_fabric_rate,2)}")
    st.info(f"Fabric Cost Per Piece: ₹ {round(fabric_cost,2)}")

    # ------------------ TRIMS ------------------
    st.subheader("Trims Cost (Per Piece)")

    trim_items = {}
    default_trims = ["Main Label", "Wash Care Label", "Tag", "Thread", "Zipper", "Mobilon Tape"]

    for trim in default_trims:
        trim_items[trim] = st.number_input(f"{trim} Cost", value=0.0)

    st.markdown("### Add Custom Trim")
    custom_trim_name = st.text_input("Trim Name")
    custom_trim_cost = st.number_input("Trim Cost", value=0.0)

    if custom_trim_name:
        trim_items[custom_trim_name] = custom_trim_cost

    trim_total = sum(trim_items.values())
    st.info(f"Total Trim Cost: ₹ {round(trim_total,2)}")

    # ------------------ PACKING ------------------
    st.subheader("Packing Cost (Per Piece)")

    packing_items = {}
    default_packing = ["Hanger", "Hang Tag", "Poly Bag", "Carton Box"]

    for item in default_packing:
        packing_items[item] = st.number_input(f"{item} Cost", value=0.0)

    st.markdown("### Add Custom Packing Item")
    custom_pack_name = st.text_input("Packing Item Name")
    custom_pack_cost = st.number_input("Packing Cost", value=0.0)

    if custom_pack_name:
        packing_items[custom_pack_name] = custom_pack_cost

    packing_total = sum(packing_items.values())
    st.info(f"Total Packing Cost: ₹ {round(packing_total,2)}")

    # ------------------ CMT ------------------
    st.subheader("CMT & Final Calculation")

    cmt = st.number_input("CMT Cost (Per Piece)", value=0.0)

    prime_cost = fabric_cost + trim_total + packing_total + cmt

    overhead_percent = st.number_input("Overhead %", value=12.0)
    margin_percent = st.number_input("Profit Margin %", value=20.0)

    overhead_value = prime_cost * (overhead_percent / 100)
    subtotal = prime_cost + overhead_value

    margin_value = subtotal * (margin_percent / 100)
    final_price = subtotal + margin_value

    st.subheader("Final Cost Summary")

    st.write(f"Prime Cost: ₹ {round(prime_cost,2)}")
    st.write(f"Overhead Amount: ₹ {round(overhead_value,2)}")
    st.write(f"Profit Amount: ₹ {round(margin_value,2)}")
    st.success(f"Final Garment Price: ₹ {round(final_price,2)}")
